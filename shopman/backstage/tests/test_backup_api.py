"""Download do cofre de dados curados — gate fino e conteúdo do arquivo.

Também prova que as entidades do backstage (de-paras, vocabulário de consumo,
salão) entram no cofre pelo registro do shop e fazem o ciclo completo — a
direção backstage → shop registrando resource é o contrato de extensão do cofre.
"""

from __future__ import annotations

import io
from io import StringIO

import pytest
from django.contrib.auth.models import Permission, User
from django.core.management import call_command
from django.urls import reverse
from openpyxl import load_workbook

from shopman.backstage.models import ConsumptionRole, ProductConsumptionTag, SeatingSpot

pytestmark = pytest.mark.django_db


def _perm(codename: str) -> Permission:
    return Permission.objects.get(codename=codename, content_type__app_label="backstage")


@pytest.fixture
def manager(db):
    user = User.objects.create_user("backup-manager", password="pw", is_staff=True)
    user.user_permissions.add(_perm("export_backup"))
    return user


@pytest.fixture
def floor_operator(db):
    user = User.objects.create_user("backup-floor", password="pw", is_staff=True)
    user.user_permissions.add(_perm("operate_production"))
    return user


def test_floor_gate_does_not_open_backup(client, floor_operator):
    client.force_login(floor_operator)
    assert client.get(reverse("api-backstage-backup-export")).status_code == 403


def test_backup_download_is_a_real_workbook(client, manager):
    SeatingSpot.objects.create(ref="mesa-1", label="Mesa 1")
    client.force_login(manager)
    response = client.get(reverse("api-backstage-backup-export"))
    assert response.status_code == 200
    assert response["Content-Disposition"].startswith('attachment; filename="backup-')
    book = load_workbook(io.BytesIO(response.content), read_only=True)
    sheets = set(book.sheetnames)
    assert {"products", "recipes", "seating_spots", "product_aliases"} <= sheets
    spots = book["seating_spots"]
    rows = list(spots.iter_rows(values_only=True))
    assert rows[0][0] == "ref"
    assert any(row[0] == "mesa-1" for row in rows[1:])


def test_backstage_entities_roundtrip_via_commands(tmp_path):
    role = ConsumptionRole.objects.create(ref="cafe", label="Café", reading="anchor")
    ProductConsumptionTag.objects.create(sku="CAFE-01", role=role)
    call_command("export_backup", "--out", str(tmp_path), stdout=StringIO())
    path = next(tmp_path.glob("backup-*.xlsx"))

    ProductConsumptionTag.objects.all().delete()
    ConsumptionRole.objects.all().update(label="Errado")

    call_command("import_backup", str(path), "--apply", stdout=StringIO())
    assert ConsumptionRole.objects.get(ref="cafe").label == "Café"
    assert ProductConsumptionTag.objects.get(sku="CAFE-01").role.ref == "cafe"
