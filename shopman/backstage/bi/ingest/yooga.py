"""Yooga → HistoricalSale/HistoricalSaleItem, por lote (BI-PLAN F6; P0 da fundação).

O export consolidado (``yooga-consolidado.xlsx``, três abas: ``Vendas``,
``Itens``, ``Produtos``) é a única origem do passado da casa antes do Shopman.
Este módulo o faz aterrissar com três garantias que a versão anterior não dava:

- **Lote com identidade.** Um ``ImportBatch`` por arquivo (nome + sha256 +
  contagens). O mesmo arquivo não entra duas vezes (``AlreadyImported``); um
  arquivo que falha fica registrado como ``failed`` com o motivo.
- **Validação na fronteira.** Cada linha passa por um modelo (pydantic) ANTES
  de virar registro: coluna faltando, tipo errado ou dinheiro ilegível é
  ``InvalidExport`` com aba e número da linha — e nada é gravado.
- **Uma transação.** Vendas e itens entram juntos ou não entram. Não existe
  mais "cabeçalho sem item" mascarado pela regra de completar depois.

Idempotente e completável continuam valendo: a chave natural é o ``pedido``
(``external_id``); rodar um export NOVO insere o que falta, completa os
metadados que faltavam nas vendas antigas (nunca sobrescreve o que já há) e
grava itens só das vendas que ainda não têm. ``--rebuild`` apaga tudo da
origem — vendas, itens e lotes — e recarrega.

⚠️ O telefone do cliente não é guardado em claro: entra como hash do E.164
(``normalize_phone``, o mesmo do guestman, para o join futuro por cliente) mais
os quatro últimos dígitos para conferência humana. Hash de telefone é
pseudonimização, não anonimato — o espaço de números é pequeno; protege da
leitura casual, não de força bruta. Não prometa mais do que isto.
"""

from __future__ import annotations

import hashlib
import logging
from collections.abc import Callable, Iterator
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Annotated, Any

from django.db import transaction
from django.utils import timezone
from pydantic import BaseModel, BeforeValidator, ConfigDict, ValidationError
from shopman.utils.phone import normalize_phone

from . import AlreadyImported, InvalidExport, sha256_of

logger = logging.getLogger(__name__)

SOURCE = "yooga"
BATCH_SIZE = 1000

# Colunas que o importador LÊ. As demais do export são ignoradas de propósito
# (``ano``, ``mes``, ``dia_semana``, ``hora_cheia`` são deriváveis da data).
REQUIRED_COLUMNS = {
    "Vendas": (
        "pedido", "data", "hora", "valor", "desconto", "acrescimo",
        "formas_pagamento", "taxa_pagamento", "operador", "cliente_id", "cliente",
        "telefone", "modalidade", "endereco", "bairro", "mesa", "observacao",
        "origem", "nfce_id",
    ),
    "Itens": (
        "pedido", "produto", "sku", "categoria", "quantidade", "valor_unitario",
        "desconto_item", "total_item",
    ),
    "Produtos": ("produto", "sku", "categoria"),
}


# ── Conversores de célula ─────────────────────────────────────────────────────
# O xlsx entrega o que quiser: float onde era reais, datetime onde era hora,
# None onde era vazio. Cada conversor aceita o que o export costuma mandar e
# levanta ValueError (que o pydantic embrulha com nome do campo) no resto.


def _cents(value: Any) -> int:
    """Reais (float/str/None) → centavos, via Decimal para não herdar erro de float."""
    if value in (None, ""):
        return 0
    try:
        return int((Decimal(str(value)) * 100).quantize(Decimal("1")))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"valor monetário ilegível: {value!r}") from exc


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        return date.fromisoformat(value.strip())
    raise ValueError(f"data ilegível: {value!r}")


def _time(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        return time.fromisoformat(value.strip())
    raise ValueError(f"hora ilegível: {value!r}")


def _text(limit: int) -> Callable[[Any], str]:
    """Texto aparado no tamanho da coluna do model; vazio/None viram ''."""

    def convert(value: Any) -> str:
        return str(value).strip()[:limit] if value not in (None, "") else ""

    return convert


def _optional_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def _quantity(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"quantidade ilegível: {value!r}") from exc


Money = Annotated[int, BeforeValidator(_cents)]
Text32 = Annotated[str, BeforeValidator(_text(32))]
Text64 = Annotated[str, BeforeValidator(_text(64))]
Text100 = Annotated[str, BeforeValidator(_text(100))]
Text200 = Annotated[str, BeforeValidator(_text(200))]
FreeText = Annotated[str, BeforeValidator(_text(500))]


class _Row(BaseModel):
    model_config = ConfigDict(extra="ignore", frozen=True)


class SaleRow(_Row):
    pedido: int
    data: Annotated[date, BeforeValidator(_date)]
    hora: Annotated[time | None, BeforeValidator(_time)] = None
    valor: Money = 0
    desconto: Money = 0
    acrescimo: Money = 0
    formas_pagamento: Text100 = ""
    taxa_pagamento: Money = 0
    operador: Text100 = ""
    cliente_id: Annotated[int | None, BeforeValidator(_optional_int)] = None
    cliente: Text200 = ""
    telefone: FreeText = ""
    modalidade: Text32 = ""
    endereco: FreeText = ""
    bairro: Text100 = ""
    mesa: Text32 = ""
    observacao: FreeText = ""
    origem: Text32 = ""
    nfce_id: Annotated[int | None, BeforeValidator(_optional_int)] = None

    @property
    def occurred_at(self) -> datetime:
        return timezone.make_aware(datetime.combine(self.data, self.hora or time.min))

    @property
    def is_delivery(self) -> bool:
        # Único rótulo confiável do histórico: origem DELIVERY ou forma de
        # pagamento "Delivery -"/"IFOOD". Mesa/balcão do Yooga nunca viram
        # verdade de canal (decisão registrada no BI-PLAN §7).
        payment = self.formas_pagamento.upper()
        return self.origem.upper() == "DELIVERY" or "DELIVERY" in payment or "IFOOD" in payment

    def metadata(self) -> dict:
        """O que o export traz e nenhuma coluna guarda. Chaves em data-schemas.md."""
        data: dict = {}
        if self.nfce_id:
            data["nfce_id"] = self.nfce_id
        phone = normalize_phone(self.telefone) if self.telefone else ""
        if phone:
            data["phone_hash"] = hashlib.sha256(phone.encode()).hexdigest()
            data["phone_last4"] = phone[-4:]
        if self.bairro:
            data["neighborhood"] = self.bairro
        if self.endereco:
            data["address"] = self.endereco
        if self.taxa_pagamento:
            data["payment_fee_q"] = self.taxa_pagamento
        if self.observacao:
            data["note"] = self.observacao
        return data


class ItemRow(_Row):
    pedido: int
    produto: Text200 = ""
    sku: Text64 = ""
    categoria: Text100 = ""
    quantidade: Annotated[Decimal, BeforeValidator(_quantity)] = Decimal("0")
    valor_unitario: Money = 0
    desconto_item: Money = 0
    total_item: Money = 0


class ProductRow(_Row):
    produto: Text200 = ""
    sku: Text64 = ""
    categoria: Text100 = ""


# ── Leitura das abas ──────────────────────────────────────────────────────────


def _rows(sheet, sheet_name: str, model: type[_Row], key: str) -> Iterator[tuple[int, _Row]]:
    """(número da linha no xlsx, linha validada) — pula linhas sem chave (rabo em branco).

    Cabeçalho é conferido antes da primeira linha: coluna que o importador lê e
    o arquivo não tem é erro do arquivo, não do meio do lote.
    """
    iterator = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(iterator)]
    except StopIteration as exc:
        raise InvalidExport(f"aba '{sheet_name}' está vazia.") from exc
    missing = [name for name in REQUIRED_COLUMNS[sheet_name] if name not in header]
    if missing:
        raise InvalidExport(
            f"aba '{sheet_name}' sem as colunas {', '.join(missing)} — arquivo inesperado."
        )
    index = {name: header.index(name) for name in REQUIRED_COLUMNS[sheet_name]}
    for line, row in enumerate(iterator, start=2):
        if row is None or index[key] >= len(row) or row[index[key]] in (None, ""):
            continue
        payload = {name: (row[at] if at < len(row) else None) for name, at in index.items()}
        try:
            yield line, model.model_validate(payload)
        except ValidationError as exc:
            problems = "; ".join(
                f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}"
                for error in exc.errors()
            )
            raise InvalidExport(f"aba '{sheet_name}', linha {line}: {problems}") from exc


# ── O importador ──────────────────────────────────────────────────────────────


def ingest(
    path: str | Path,
    *,
    rebuild: bool = False,
    imported_by=None,
    log: Callable[[str], None] = lambda message: None,
):
    """Importa o export em UM lote e devolve o ``ImportBatch`` concluído.

    Levanta ``AlreadyImported`` (mesmo hash já concluído nesta origem, sem
    ``rebuild``) ou ``InvalidExport`` (arquivo/aba/coluna/linha inválida). Em
    qualquer falha depois de o arquivo abrir, um lote ``failed`` fica gravado
    com o motivo — a transação de dados desfaz tudo, o registro da falha não.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependência declarada no pyproject
        raise InvalidExport("openpyxl não instalado (pip install openpyxl).") from exc

    from shopman.backstage.models import HistoricalSale, HistoricalSaleItem, ImportBatch

    path = Path(path)
    if not path.is_file():
        raise InvalidExport(f"Arquivo não encontrado: {path}")
    digest = sha256_of(path)

    if not rebuild:
        done = ImportBatch.objects.filter(
            source=SOURCE, file_sha256=digest, status=ImportBatch.Status.DONE
        ).first()
        if done is not None:
            raise AlreadyImported(
                f"Este arquivo já foi importado em {timezone.localtime(done.imported_at):%d/%m/%Y %H:%M} "
                f"(lote #{done.pk}). Para recarregar do zero use --rebuild."
            )

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in REQUIRED_COLUMNS:
            if sheet not in workbook.sheetnames:
                raise InvalidExport(f"Aba '{sheet}' ausente no export — arquivo inesperado.")

        try:
            with transaction.atomic():
                if rebuild:
                    deleted, _ = HistoricalSale.objects.filter(source=SOURCE).delete()
                    ImportBatch.objects.filter(source=SOURCE).delete()
                    log(f"--rebuild: {deleted} registros de {SOURCE} removidos.")

                batch = ImportBatch.objects.create(
                    source=SOURCE,
                    file_name=path.name[:200],
                    file_sha256=digest,
                    imported_by=imported_by,
                    status=ImportBatch.Status.DONE,
                )
                product_map = _product_map(workbook["Produtos"])
                summary = _ingest_sales(workbook["Vendas"], batch, HistoricalSale)
                summary["items_created"] = _ingest_items(
                    workbook["Itens"], HistoricalSale, HistoricalSaleItem, product_map
                )
                for field, value in summary.items():
                    setattr(batch, field, value)
                batch.save(update_fields=list(summary))
        except Exception as exc:
            # A transação desfez os dados; o registro da falha fica, e o erro sobe.
            _record_failure(ImportBatch, path, digest, imported_by, exc)
            raise
    finally:
        workbook.close()

    # O passado mudou; o passado materializado tem de mudar junto — do zero,
    # porque um export pode trazer dias que a série ainda não conhecia.
    from shopman.backstage.bi.daily_series import refresh_all

    days = refresh_all()
    log(f"↻ série diária do B.I. recomputada: {days} dias.")

    total = HistoricalSale.objects.filter(source=SOURCE).count()
    log(
        f"✅ lote #{batch.pk}: {batch.sales_created} vendas novas, "
        f"{batch.sales_skipped} já existiam ({batch.sales_completed} completadas), "
        f"{batch.items_created} itens gravados. Total {SOURCE}: {total}."
    )
    return batch


def _record_failure(batch_model, path: Path, digest: str, imported_by, exc: Exception) -> None:
    """Lote ``failed`` com o motivo: importação que falha em silêncio é o pior caso."""
    logger.warning("bi.ingest.yooga.failed file=%s error=%s", path.name, exc)
    batch_model.objects.create(
        source=SOURCE,
        file_name=path.name[:200],
        file_sha256=digest,
        imported_by=imported_by,
        status=batch_model.Status.FAILED,
        error=str(exc)[:2000],
    )


def _product_map(sheet) -> dict[str, tuple[str, str]]:
    """nome (minúsculo) → (sku, categoria) do catálogo do export, para linha antiga sem sku."""
    mapping: dict[str, tuple[str, str]] = {}
    for _line, row in _rows(sheet, "Produtos", ProductRow, key="produto"):
        mapping[row.produto.lower()] = (row.sku, row.categoria)
    return mapping


def _ingest_sales(sheet, batch, sale_model) -> dict[str, int]:
    """Grava as vendas do export; devolve as contagens no vocabulário do ``ImportBatch``."""
    # Uma passada pelo que já existe: chave natural → (pk, metadata). Cabe na
    # memória (81k vendas) e evita um SELECT por linha do export.
    existing: dict[int, tuple[int | None, dict]] = {
        external_id: (pk, metadata or {})
        for external_id, pk, metadata in sale_model.objects.filter(source=SOURCE).values_list(
            "external_id", "id", "metadata"
        )
    }
    summary = {"rows_read": 0, "sales_created": 0, "sales_skipped": 0, "sales_completed": 0}
    to_create: list = []
    to_complete: list = []

    def flush():
        if to_create:
            sale_model.objects.bulk_create(to_create)
            summary["sales_created"] += len(to_create)
            to_create.clear()
        if to_complete:
            sale_model.objects.bulk_update(to_complete, ["metadata"])
            summary["sales_completed"] += len(to_complete)
            to_complete.clear()

    for _line, row in _rows(sheet, "Vendas", SaleRow, key="pedido"):
        summary["rows_read"] += 1
        metadata = row.metadata()
        known = existing.get(row.pedido)
        if known is not None:
            summary["sales_skipped"] += 1
            pk, stored = known
            if pk is None:
                continue  # pedido repetido dentro do próprio arquivo: a primeira linha vale
            # Completa o que faltava, nunca sobrescreve: a primeira leitura de
            # um dado é a que vale; um export posterior só preenche lacunas.
            gained = {key: value for key, value in metadata.items() if key not in stored}
            if gained:
                merged = {**stored, **gained}
                existing[row.pedido] = (pk, merged)
                to_complete.append(sale_model(pk=pk, metadata=merged))
            if len(to_complete) >= BATCH_SIZE:
                flush()
            continue
        sale = sale_model(
            source=SOURCE,
            batch=batch,
            external_id=row.pedido,
            occurred_at=row.occurred_at,
            total_q=row.valor,
            discount_q=row.desconto,
            surcharge_q=row.acrescimo,
            payment=row.formas_pagamento,
            operator=row.operador,
            modality=row.modalidade,
            origin=row.origem,
            table_label=row.mesa,
            is_delivery=row.is_delivery,
            customer_external_id=row.cliente_id,
            customer_name=row.cliente,
            metadata=metadata,
        )
        existing[row.pedido] = (None, metadata)  # visto neste lote; pk só existe após o flush
        to_create.append(sale)
        if len(to_create) >= BATCH_SIZE:
            flush()
    flush()
    return summary


def _ingest_items(sheet, sale_model, item_model, product_map) -> int:
    # Só grava itens de vendas que ainda não os têm — completável: se um
    # export futuro trouxer itens de vendas antigas, eles entram sem duplicar.
    sale_pk = dict(sale_model.objects.filter(source=SOURCE).values_list("external_id", "id"))
    with_items = set(
        item_model.objects.filter(sale__source=SOURCE)
        .values_list("sale__external_id", flat=True)
        .distinct()
    )
    seq_by_sale: dict[int, int] = {}
    written = 0
    pending: list = []
    for _line, row in _rows(sheet, "Itens", ItemRow, key="pedido"):
        if row.pedido in with_items or row.pedido not in sale_pk:
            continue
        sku, category = row.sku, row.categoria
        if not sku or not category:
            mapped_sku, mapped_category = product_map.get(row.produto.lower(), ("", ""))
            sku = sku or mapped_sku
            category = category or mapped_category
        seq = seq_by_sale.get(row.pedido, 0) + 1
        seq_by_sale[row.pedido] = seq
        pending.append(
            item_model(
                sale_id=sale_pk[row.pedido],
                seq=seq,
                product_name=row.produto,
                sku=sku,
                category=category,
                qty=row.quantidade,
                unit_price_q=row.valor_unitario,
                discount_q=row.desconto_item,
                line_total_q=row.total_item,
            )
        )
        if len(pending) >= BATCH_SIZE:
            item_model.objects.bulk_create(pending)
            written += len(pending)
            pending = []
    if pending:
        item_model.objects.bulk_create(pending)
        written += len(pending)
    return written
