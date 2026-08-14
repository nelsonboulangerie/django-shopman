"""Ingere o export consolidado do Yooga em HistoricalSale/Item (BI-PLAN F6).

Idempotente e completável por construção: a chave natural é o ``pedido`` do
Yooga (``external_id``); rodar de novo só insere o que falta. ``--rebuild``
apaga tudo de ``source=yooga`` e recarrega — o histórico é recomputável a
partir do arquivo, que é a origem (o original no Drive nunca é editado; este
comando abre o xlsx em modo somente leitura).

O arquivo NÃO entra no git (``var/`` está no .gitignore). Abas esperadas:
``Vendas`` (cabeçalho da venda), ``Itens`` (linhas), ``Produtos`` (mapa
nome→sku/categoria para linhas antigas sem sku).
"""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

SOURCE = "yooga"
BATCH = 1000


def _q(value) -> int:
    """Reais (float/str/None do xlsx) → centavos int, sem erro de float."""
    if value in (None, ""):
        return 0
    try:
        return int((Decimal(str(value)) * 100).quantize(Decimal("1")))
    except (InvalidOperation, ValueError) as exc:
        raise CommandError(f"Valor monetário inválido no export: {value!r}") from exc


def _dt(raw_date, raw_time):
    """data + hora do export (str ou objetos) → datetime local aware."""
    if isinstance(raw_date, datetime):
        raw_date = raw_date.date()
    if isinstance(raw_date, str):
        raw_date = date.fromisoformat(raw_date.strip())
    if raw_time in (None, ""):
        raw_time = time.min
    elif isinstance(raw_time, datetime):
        raw_time = raw_time.time()
    elif isinstance(raw_time, str):
        raw_time = time.fromisoformat(raw_time.strip())
    return timezone.make_aware(datetime.combine(raw_date, raw_time))


def _text(value, limit: int) -> str:
    return str(value).strip()[:limit] if value not in (None, "") else ""


def _is_delivery(origin: str, payment: str) -> bool:
    # Único rótulo confiável do histórico: origem DELIVERY ou forma de
    # pagamento "Delivery -"/"IFOOD" (decisão registrada na memória do projeto;
    # mesa/balcão do Yooga nunca viram verdade de canal).
    upper_payment = payment.upper()
    return origin.upper() == "DELIVERY" or "DELIVERY" in upper_payment or "IFOOD" in upper_payment


class Command(BaseCommand):
    help = "Ingere o export consolidado do Yooga (xlsx) em HistoricalSale/HistoricalSaleItem."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho do yooga-consolidado.xlsx.")
        parser.add_argument(
            "--rebuild",
            action="store_true",
            help="Apaga tudo de source=yooga e recarrega do arquivo (recomputável por design).",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl não instalado (pip install openpyxl).") from exc

        from shopman.backstage.models import HistoricalSale, HistoricalSaleItem

        try:
            workbook = openpyxl.load_workbook(options["file"], read_only=True, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"Arquivo não encontrado: {options['file']}") from exc

        for sheet in ("Vendas", "Itens", "Produtos"):
            if sheet not in workbook.sheetnames:
                raise CommandError(f"Aba '{sheet}' ausente no export — arquivo inesperado.")

        if options["rebuild"]:
            deleted, _ = HistoricalSale.objects.filter(source=SOURCE).delete()
            self.stdout.write(f"--rebuild: {deleted} registros de {SOURCE} removidos.")

        product_map = self._product_map(workbook["Produtos"])
        existing = set(
            HistoricalSale.objects.filter(source=SOURCE).values_list("external_id", flat=True)
        )
        preexisting = len(existing)

        created = self._ingest_sales(workbook["Vendas"], HistoricalSale, existing)
        items = self._ingest_items(
            workbook["Itens"], HistoricalSale, HistoricalSaleItem, product_map
        )
        workbook.close()

        total = HistoricalSale.objects.filter(source=SOURCE).count()
        self.stdout.write(
            f"✅ {created} vendas novas ({preexisting} já existiam), "
            f"{items} itens gravados. Total {SOURCE}: {total}."
        )

    def _product_map(self, sheet) -> dict[str, tuple[str, str]]:
        rows = sheet.iter_rows(values_only=True)
        header = [str(cell or "").strip() for cell in next(rows)]
        col = {name: index for index, name in enumerate(header)}
        mapping: dict[str, tuple[str, str]] = {}
        for row in rows:
            if row is None or row[col["produto"]] in (None, ""):
                continue
            name = str(row[col["produto"]]).strip()
            mapping[name.lower()] = (
                _text(row[col["sku"]], 64),
                _text(row[col["categoria"]], 100),
            )
        return mapping

    def _ingest_sales(self, sheet, sale_model, existing: set[int]) -> int:
        rows = sheet.iter_rows(values_only=True)
        header = [str(cell or "").strip() for cell in next(rows)]
        col = {name: index for index, name in enumerate(header)}
        created = 0
        batch: list = []
        for row in rows:
            if row is None or row[col["pedido"]] in (None, ""):
                continue
            external_id = int(row[col["pedido"]])
            if external_id in existing:
                continue
            existing.add(external_id)
            origin = _text(row[col["origem"]], 32)
            payment = _text(row[col["formas_pagamento"]], 100)
            metadata = {}
            nfce = row[col["nfce_id"]]
            if nfce not in (None, "", 0):
                metadata["nfce_id"] = nfce
            batch.append(
                sale_model(
                    source=SOURCE,
                    external_id=external_id,
                    occurred_at=_dt(row[col["data"]], row[col["hora"]]),
                    total_q=_q(row[col["valor"]]),
                    discount_q=_q(row[col["desconto"]]),
                    surcharge_q=_q(row[col["acrescimo"]]),
                    payment=payment,
                    operator=_text(row[col["operador"]], 100),
                    modality=_text(row[col["modalidade"]], 32),
                    origin=origin,
                    table_label=_text(row[col["mesa"]], 32),
                    is_delivery=_is_delivery(origin, payment),
                    customer_external_id=row[col["cliente_id"]] or None,
                    customer_name=_text(row[col["cliente"]], 200),
                    metadata=metadata,
                )
            )
            if len(batch) >= BATCH:
                sale_model.objects.bulk_create(batch)
                created += len(batch)
                batch = []
        if batch:
            sale_model.objects.bulk_create(batch)
            created += len(batch)
        return created

    def _ingest_items(self, sheet, sale_model, item_model, product_map) -> int:
        # Só grava itens de vendas que ainda não os têm — completável: se um
        # export futuro trouxer itens de vendas antigas, eles entram sem duplicar.
        sale_pk = dict(
            sale_model.objects.filter(source=SOURCE).values_list("external_id", "id")
        )
        with_items = set(
            item_model.objects.filter(sale__source=SOURCE)
            .values_list("sale__external_id", flat=True)
            .distinct()
        )

        rows = sheet.iter_rows(values_only=True)
        header = [str(cell or "").strip() for cell in next(rows)]
        col = {name: index for index, name in enumerate(header)}

        seq_by_sale: dict[int, int] = {}
        written = 0
        batch: list = []
        with transaction.atomic():
            for row in rows:
                if row is None or row[col["pedido"]] in (None, ""):
                    continue
                external_id = int(row[col["pedido"]])
                if external_id in with_items or external_id not in sale_pk:
                    continue
                name = _text(row[col["produto"]], 200)
                sku = _text(row[col["sku"]], 64)
                category = _text(row[col["categoria"]], 100)
                if not sku or not category:
                    mapped_sku, mapped_category = product_map.get(name.lower(), ("", ""))
                    sku = sku or mapped_sku
                    category = category or mapped_category
                seq = seq_by_sale.get(external_id, 0) + 1
                seq_by_sale[external_id] = seq
                batch.append(
                    item_model(
                        sale_id=sale_pk[external_id],
                        seq=seq,
                        product_name=name,
                        sku=sku,
                        category=category,
                        qty=Decimal(str(row[col["quantidade"]] or 0)),
                        unit_price_q=_q(row[col["valor_unitario"]]),
                        discount_q=_q(row[col["desconto_item"]]),
                        line_total_q=_q(row[col["total_item"]]),
                    )
                )
                if len(batch) >= BATCH:
                    item_model.objects.bulk_create(batch)
                    written += len(batch)
                    batch = []
            if batch:
                item_model.objects.bulk_create(batch)
                written += len(batch)
        return written
