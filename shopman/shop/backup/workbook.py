"""Leitura e escrita do arquivo do cofre — XLSX (uma aba por entidade) ou CSVs.

A escrita grava tudo como TEXTO, na forma que os widgets do import-export
renderizam — a mesma forma que eles sabem ler de volta. A leitura normaliza o
que o Excel/Google Sheets tiver convertido no meio do caminho (número, booleano,
data viram célula tipada quando alguém edita a planilha) de volta para texto no
formato dos widgets. É isso que torna o ciclo exportar → editar no Sheets →
importar um ciclo fechado.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
import re
from pathlib import Path

import tablib
from openpyxl import Workbook, load_workbook

from shopman.shop.backup import registry


#: Primeiro caractere que faz Excel/Sheets tratar o texto como fórmula ativa.
_FORMULA_CHARS = ("=", "@", "\t", "\r")
_NUMERIC_RE = re.compile(r"^[+-]?\d+([.,]\d+)?$")


def _needs_escape(text: str) -> bool:
    """Texto que uma planilha interpretaria como fórmula, não como dado.

    ``+``/``-`` só são perigosos quando NÃO formam um número puro: "-10" é o
    número -10 em qualquer planilha; "-2+cmd()" é fórmula. Escapar o número
    sujaria a célula sem ganhar nada.
    """
    if not text:
        return False
    if text[0] in _FORMULA_CHARS:
        return True
    return text[0] in "+-" and not _NUMERIC_RE.match(text)


def _escape_cell(text: str) -> str:
    """Prefixa ``'`` no que viraria fórmula ao abrir no Sheets/Excel.

    Sem isto, um nome curado começando com ``=`` executa ao abrir a planilha
    viva E — pior — volta do import como o VALOR COMPUTADO (o leitor usa
    ``data_only``), corrompendo o dado no restore. Também escapa ``'`` + texto
    perigoso, para a des-escapada da leitura ser uma bijeção exata.
    """
    if _needs_escape(text) or (text.startswith("'") and _needs_escape(text[1:])):
        return "'" + text
    return text


def _unescape_cell(text: str) -> str:
    if text.startswith("'") and (_needs_escape(text[1:]) or (
        text[1:].startswith("'") and _needs_escape(text[2:])
    )):
        return text[1:]
    return text


def _cell_to_text(value) -> str:
    """Devolve a célula à forma textual que os widgets do import-export leem."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def export_datasets(*, with_read_only: bool = False) -> dict[str, tablib.Dataset]:
    """Exporta cada entidade registrada para um Dataset, na ordem do registro.

    As abas somente-leitura (transacionais, para conferência) só entram quando
    pedidas — o arquivo padrão é o cofre curado, que faz round-trip.
    """
    out: dict[str, tablib.Dataset] = {}
    for entry in registry.entries():
        if entry.read_only and not with_read_only:
            continue
        out[entry.name] = entry.resource_class().export()
    return out


def write_xlsx(datasets: dict[str, tablib.Dataset]) -> bytes:
    book = Workbook()
    book.remove(book.active)
    for name, dataset in datasets.items():
        sheet = book.create_sheet(title=name)
        sheet.append(list(dataset.headers or []))
        for row in dataset:
            sheet.append([_escape_cell("" if v is None else str(v)) for v in row])
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def read_xlsx(path: Path) -> dict[str, tablib.Dataset]:
    """Lê o arquivo de volta: uma aba = um Dataset, células normalizadas p/ texto."""
    book = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, tablib.Dataset] = {}
    for sheet in book.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        dataset = tablib.Dataset()
        dataset.headers = [str(h) for h in headers if h is not None] if headers else []
        width = len(dataset.headers)
        for row in rows:
            values = [_unescape_cell(_cell_to_text(v)) for v in row[:width]]
            values += [""] * (width - len(values))
            if any(v != "" for v in values):
                dataset.append(values)
        out[sheet.title] = dataset
    book.close()
    return out


def write_csv_dir(datasets: dict[str, tablib.Dataset], out_dir: Path) -> list[Path]:
    """Um CSV por entidade — a forma que dá diff legível em git."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    for name, dataset in datasets.items():
        path = out_dir / f"{name}.csv"
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(dataset.headers or [])
            for row in dataset:
                writer.writerow([_escape_cell("" if v is None else str(v)) for v in row])
        written.append(path)
    return written


def read_csv_dir(dir_path: Path) -> dict[str, tablib.Dataset]:
    out: dict[str, tablib.Dataset] = {}
    for path in sorted(dir_path.glob("*.csv")):
        with path.open("r", newline="", encoding="utf-8") as fh:
            rows = list(csv.reader(fh))
        dataset = tablib.Dataset()
        dataset.headers = rows[0] if rows else []
        for row in rows[1:]:
            if any(v != "" for v in row):
                dataset.append([_unescape_cell(v) for v in row])
        out[path.stem] = dataset
    return out
